#!/usr/bin/env python3
"""Build the colored Base.xml worksheet (docs/base_xml_worksheet.xlsx).

Categories per item, engine-derived, with a distinct colour each. Run from the
pics_tool package root:  python3 tools/build_worksheet.py
"""
import sys
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pics_tool.generate.mcore_engine import (compute_mcore_pics, _TRANSPORT_SEEDS,
    _ROLE_SEEDS, _ONBOARDING_SEEDS, _BDX_REQUESTOR, _BDX_PROVIDER)
from pics_tool.generate.profile import DeviceProfile
from pics_tool.generate.template_io import base_template_path

REF_PATH = "extended-color-light-manually-scaped-pics/Base.xml"
root = ET.parse(str(base_template_path("1.6"))).getroot()
feat = {}; conds = {}; order = []
for pi in root.iter("picsItem"):
    n = (pi.findtext("itemNumber") or "").strip()
    if not n:
        continue
    order.append(n); feat[n] = " ".join((pi.findtext("feature") or "").split())
    conds[n] = " ; ".join([f"{(s.text or '').strip()} if {(s.attrib.get('cond','') or '').strip()}"
                           if (s.attrib.get('cond','') or '').strip() else (s.text or '').strip()
                           for s in pi.findall("status")])
I = set(order)


def en(role="commissionee", transports=("wifi_2g",), ble=None,
       onboarding=("qr", "manual_pairing_code"), clusters=frozenset(), icd=False):
    d = {"spec_version": "1.6", "device_type": "On/Off Light", "transport": list(transports),
         "role": role, "onboarding": list(onboarding), "is_icd": icd}
    if ble is not None:
        d["ble_commissioning"] = ble
    return compute_mcore_pics(DeviceProfile.from_dict(d), "1.6", set(clusters)) & I


DIMS = {"transport": [dict(transports=(t,)) for t in ("wifi_2g","wifi_5g","thread","ethernet")],
        "ble": [dict(ble=True), dict(ble=False)],
        "role": [dict(role=r) for r in ("commissionee","commissioner","controller")],
        "onboarding": [dict(onboarding=x) for x in ((),("qr",),("manual_pairing_code",),("nfc",))],
        "device_types": [dict(clusters=c) for c in (frozenset(),frozenset({"0x002a"}),
                         frozenset({"0x0029"}),frozenset({"0x0039","0x0751"}))],
        "is_icd": [dict(icd=False), dict(icd=True)]}
dec = defaultdict(set); reach = set()
for dim, vs in DIMS.items():
    ss = [en(**v) for v in vs]
    for s in ss:
        reach |= s
    for it in set().union(*ss) - set.intersection(*map(set, ss)):
        dec[it].add(dim)
for r in ("commissionee","commissioner","controller"):
    for cl in (frozenset(),frozenset({"0x002a"}),frozenset({"0x0029"})):
        reach |= en(role=r, clusters=cl)
manual = I - reach
seeds = set().union(*_TRANSPORT_SEEDS.values(),*_ROLE_SEEDS.values(),*_ONBOARDING_SEEDS.values(),{"MCORE.COM.BLE"})
comp = {"MCORE.OTA.Requestor","MCORE.OTA.Provider","MCORE.BRIDGE"} | _BDX_REQUESTOR | _BDX_PROVIDER
ref = {}
for pi in ET.parse(REF_PATH).getroot().iter("picsItem"):
    n = (pi.findtext("itemNumber") or "").strip()
    if n:
        ref[n] = (pi.findtext("support") or "").strip().lower() == "true"


def cat(n):
    if n == "MCORE.ROLE.COMMISSIONEE":
        return "default_true"                 # universal — every DUT is a commissionee
    if n.startswith("MCORE.IDM.C"):
        return "client"
    if n.startswith("MCORE.IDM.S"):
        return "server"
    if n in manual:
        return "manual_ref_true" if ref.get(n) else "manual"
    if n in comp:
        return "composition"
    if n in seeds:
        return "input"
    if " if " in conds.get(n, ""):
        return "derived"
    return "input"


def should(n, c):
    if c in ("default_true", "manual_ref_true"):
        return "true"
    if c == "server":
        return "true" if n == "MCORE.IDM.S" else "false"
    if c in ("client", "manual"):
        return "false"
    return "auto"                              # input / derived / composition


def note(n, c, dims):
    d = "/".join(dims) or "profile"
    if c == "default_true":
        return "universal — every DUT is a commissionee (default true)"
    if c == "client":
        return "IM CLIENT capability — on only if the device is an IM client"
    if c == "server":
        return ("IM server — true for a device that hosts clusters" if n == "MCORE.IDM.S"
                else "optional IM server capability — product-specific")
    if c == "manual_ref_true":
        return "MANUAL, but set TRUE in the curated reference PICS — review/keep"
    if c == "manual":
        return "manual — optional/product-specific; off by default"
    if c == "input":
        return f"auto — set by the {d} input"
    if c == "derived":
        return f"auto — derived (mandatory) via condition from {d}"
    if c == "composition":
        return "auto — from node_device_types (OTA/bridge cluster present)"
    return ""


COLORS = {"default_true": "D7F0F0", "input": "DDEBF7", "derived": "E7DDF7", "composition": "DDF7E3",
          "client": "FCE4D6", "server": "E2EFDA", "manual": "F7E0DD", "manual_ref_true": "FFF2CC"}

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Base.xml"
headers = ["item", "group", "category", "should_enable", "decided_by", "conformance",
           "ref_extended_color_light", "description", "notes"]
ws.append(headers)
hf = PatternFill("solid", fgColor="404040")
for j in range(1, len(headers) + 1):
    c = ws.cell(1, j); c.fill = hf; c.font = Font(bold=True, color="FFFFFF")
for n in order:
    c = cat(n); dims = sorted(dec.get(n, set()))
    ws.append([n, ".".join(n.split(".")[:2]), c, should(n, c),
               ", ".join(dims), conds[n], str(ref.get(n, "")).lower() if n in ref else "",
               feat[n], note(n, c, dims)])
    fill = PatternFill("solid", fgColor=COLORS[c])
    for j in range(1, len(headers) + 1):
        ws.cell(ws.max_row, j).fill = fill
ws.freeze_panes = "A2"
for j, wd in enumerate([34, 16, 16, 13, 22, 30, 12, 56, 52], 1):
    ws.column_dimensions[get_column_letter(j)].width = wd
for r in range(1, ws.max_row + 1):
    ws.cell(r, 8).alignment = Alignment(wrap_text=True, vertical="top")
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
wb.save("docs/base_xml_worksheet.xlsx")
print("categories:", dict(Counter(cat(n) for n in order)))
